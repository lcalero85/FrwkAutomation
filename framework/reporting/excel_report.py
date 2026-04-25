from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from framework.reporting.models import ExecutionSummary


class ExcelReportGenerator:
    def __init__(self, output_dir: str | Path = "reports/excel") -> None:
        self.root_path = Path(__file__).resolve().parents[2]
        self.output_dir = self.root_path / output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, summary: ExecutionSummary) -> Path:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        self._build_summary(ws_summary, summary)
        self._build_steps(wb.create_sheet("Pasos ejecutados"), summary)
        self._build_errors(wb.create_sheet("Errores"), summary)
        self._build_metrics(wb.create_sheet("Métricas"), summary)
        output_path = self.output_dir / f"execution_report_{summary.execution_id}.xlsx"
        wb.save(output_path)
        return output_path

    def _title_style(self, cell):
        cell.font = Font(bold=True, size=16, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    def _header_style(self, row):
        fill = PatternFill("solid", fgColor="D9EAF7")
        border = Border(bottom=Side(style="thin", color="8EAADB"))
        for cell in row:
            cell.font = Font(bold=True, color="1F1F1F")
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _status_fill(self, status: str):
        colors = {"PASSED": "C6EFCE", "FAILED": "FFC7CE", "SKIPPED": "FFEB9C", "RUNNING": "D9EAF7"}
        return PatternFill("solid", fgColor=colors.get(status, "E7E6E6"))

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _build_summary(self, ws, summary: ExecutionSummary) -> None:
        ws.merge_cells("A1:B1")
        ws["A1"] = "AutoTest Pro Framework - Reporte Ejecutivo"
        self._title_style(ws["A1"])
        rows = [
            ("Execution ID", summary.execution_id), ("Proyecto", summary.project_name),
            ("Suite", summary.suite_name), ("Caso", summary.test_name), ("Estado", summary.status),
            ("Ambiente", summary.environment), ("URL Base", summary.base_url), ("Navegador", summary.browser),
            ("Inicio", summary.started_at), ("Fin", summary.finished_at),
            ("Duración", f"{summary.duration_seconds} segundos"), ("Total pasos", summary.total_steps),
            ("Pasos exitosos", summary.passed_steps), ("Pasos fallidos", summary.failed_steps),
            ("Pasos omitidos", summary.skipped_steps), ("Porcentaje éxito", f"{summary.success_percentage}%"),
        ]
        for index, (label, value) in enumerate(rows, start=3):
            ws.cell(row=index, column=1, value=label).font = Font(bold=True)
            ws.cell(row=index, column=2, value=value)
            if label == "Estado":
                ws.cell(row=index, column=2).fill = self._status_fill(str(value))
        rec_row = len(rows) + 5
        ws.cell(row=rec_row, column=1, value="Recomendaciones").font = Font(bold=True)
        for i, rec in enumerate(summary.recommendations, start=rec_row + 1):
            ws.cell(row=i, column=1, value=f"- {rec}")
        self._auto_width(ws)

    def _build_steps(self, ws, summary: ExecutionSummary) -> None:
        headers = ["Orden", "Nombre", "Acción", "Selector", "Estado", "Mensaje", "Screenshot", "Duración", "Error"]
        ws.append(headers)
        self._header_style(ws[1])
        for step in summary.steps:
            ws.append([step.order, step.name, step.action, step.selector, step.status, step.message, step.screenshot_path, step.duration_seconds, step.error_detail])
            ws.cell(row=ws.max_row, column=5).fill = self._status_fill(step.status)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)

    def _build_errors(self, ws, summary: ExecutionSummary) -> None:
        headers = ["Orden", "Paso", "Acción", "Selector", "Error", "Screenshot"]
        ws.append(headers)
        self._header_style(ws[1])
        for step in summary.steps:
            if step.status == "FAILED":
                ws.append([step.order, step.name, step.action, step.selector, step.error_detail, step.screenshot_path])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)

    def _build_metrics(self, ws, summary: ExecutionSummary) -> None:
        ws.append(["Estado", "Cantidad"])
        self._header_style(ws[1])
        ws.append(["PASSED", summary.passed_steps])
        ws.append(["FAILED", summary.failed_steps])
        ws.append(["SKIPPED", summary.skipped_steps])
        chart = PieChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
        chart.title = "Distribución de pasos"
        ws.add_chart(chart, "D2")
        self._auto_width(ws)
